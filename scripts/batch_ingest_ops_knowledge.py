#!/usr/bin/env python3
"""Batch ingest all documents from docs/ops-knowledge/raw/ into ops_knowledge RAG.

Directory mapping:
  raw/预案/* → doc_type=emergency
  raw/SOP/*  → doc_type=sop

Usage:
  python scripts/batch_ingest_ops_knowledge.py [--dry-run]

--dry-run: only list files, don't ingest.
"""

import hashlib
import sys
import uuid
import datetime
import argparse
from pathlib import Path

# Add mcp-servers to path
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
MCP_SERVERS = PROJECT_ROOT / "mcp-servers" / "ops-knowledge"
sys.path.insert(0, str(MCP_SERVERS))

from config import get_config
from db.metadata_client import MetadataClient
from rag.ops_knowledge_rag import OpsKnowledgeRAG, chunk_by_doc_type

# --- File readers (same logic as knowledge_upload tool) ---


def _file_hash(file_path: str) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for block in iter(lambda: f.read(8192), b""):
            h.update(block)
    return h.hexdigest()


def _read_file(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    if ext in (".md", ".txt", ".csv"):
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    elif ext == ".pdf":
        return _read_pdf(file_path)
    elif ext == ".docx":
        return _read_docx(file_path)
    elif ext == ".xlsx":
        return _read_xlsx(file_path)
    else:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()


def _read_pdf(file_path: str) -> str:
    try:
        import fitz

        doc = fitz.open(file_path)
        return "\n".join(page.get_text() for page in doc)
    except ImportError:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        print(f"  [WARN] Failed to read PDF {file_path}: {e}")
        return ""


def _read_docx(file_path: str) -> str:
    try:
        from docx import Document

        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        print(f"  [WARN] Failed to read DOCX {file_path}: {e}")
        return ""


def _read_xlsx(file_path: str) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
        return "\n".join(rows)
    except ImportError:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        print(f"  [WARN] Failed to read XLSX {file_path}: {e}")
        return ""


def classify_doc_type(rel_path: str) -> str:
    """Determine doc_type from directory structure."""
    parts = Path(rel_path).parts
    # parts[0] = "docs", parts[1] = "ops-knowledge", parts[2] = "raw", parts[3] = "预案" or "SOP"
    if len(parts) >= 4:
        top_dir = parts[3]
        if top_dir == "预案":
            return "emergency"
        elif top_dir == "SOP":
            return "sop"
    return "sop"  # default


def extract_title(filename: str) -> str:
    """Extract a clean title from filename."""
    stem = Path(filename).stem
    # Remove common prefixes like IT0001- or AQ001-
    # Just use the full stem as title, it's descriptive enough
    return stem.strip()


def main():
    parser = argparse.ArgumentParser(description="Batch ingest ops-knowledge docs")
    parser.add_argument(
        "--dry-run", action="store_true", help="Only list files, don't ingest"
    )
    args = parser.parse_args()

    cfg = get_config()
    raw_dir = (PROJECT_ROOT / cfg.chroma.raw_dir).resolve()

    if not raw_dir.exists():
        print(f"ERROR: raw_dir not found: {raw_dir}")
        sys.exit(1)

    # Collect all supported files
    supported_ext = {".pdf", ".docx", ".xlsx", ".txt", ".md", ".csv"}
    files = []
    for f in sorted(raw_dir.rglob("*")):
        if f.is_file() and f.suffix.lower() in supported_ext:
            files.append(f)

    print(f"Found {len(files)} files in {raw_dir}")

    if args.dry_run:
        for f in files:
            rel = f.relative_to(raw_dir)
            doc_type = classify_doc_type(str(f.relative_to(PROJECT_ROOT)))
            title = extract_title(f.name)
            print(f"  [{doc_type:9s}] {title}")
        print(f"\nTotal: {len(files)} files (dry-run, no ingestion)")
        return

    # Initialize RAG and DB
    print("\nInitializing RAG...")
    rag = OpsKnowledgeRAG(
        persist_dir=cfg.chroma.persist_dir,
        ollama_base_url=cfg.chroma.ollama_base_url,
        ollama_model=cfg.chroma.ollama_model,
        collection_name=cfg.chroma.collection_name,
    )
    rag.initialize()
    db = MetadataClient(db_path=cfg.sqlite.db_path)

    # Stats
    stats = {"success": 0, "skipped": 0, "error": 0, "empty": 0, "total_chunks": 0}

    for i, f in enumerate(files, 1):
        rel = f.relative_to(raw_dir)
        doc_type = classify_doc_type(str(f.relative_to(PROJECT_ROOT)))
        title = extract_title(f.name)

        print(f"\n[{i}/{len(files)}] {rel} ({doc_type})")

        # Check hash for dedup
        try:
            file_hash = _file_hash(str(f))
        except Exception as e:
            print(f"  [SKIP] Cannot read file hash: {e}")
            stats["error"] += 1
            continue

        existing = db.get_document_by_hash(file_hash)
        if existing:
            print(f"  [SKIP] Hash unchanged (doc_id={existing['doc_id'][:8]}...)")
            stats["skipped"] += 1
            continue

        # Check title dedup
        existing_by_title = db.get_document_by_title(title)
        if existing_by_title:
            old_doc_id = existing_by_title["doc_id"]
            deleted = rag.delete_chunks_by_doc_id(old_doc_id)
            db.delete_document(old_doc_id)
            print(f"  [OVERWRITE] Title exists, deleted {deleted} old chunks")

        # Read file
        content = _read_file(str(f))
        if not content.strip():
            print(f"  [SKIP] Empty content")
            stats["empty"] += 1
            continue

        # Chunk
        raw_chunks = chunk_by_doc_type(content, doc_type)
        if not raw_chunks:
            print(f"  [SKIP] No valid chunks after splitting")
            stats["empty"] += 1
            continue

        # Build chunks with metadata
        doc_id = str(uuid.uuid4())
        upload_date = datetime.datetime.now().isoformat()
        source_rel = str(f.relative_to(PROJECT_ROOT))

        chunks = []
        for idx, chunk_content in enumerate(raw_chunks):
            chunks.append(
                {
                    "content": chunk_content,
                    "metadata": {
                        "doc_id": doc_id,
                        "doc_type": doc_type,
                        "device_vendor": "",
                        "device_type": "",
                        "source_file": source_rel,
                        "chunk_index": idx,
                        "total_chunks": len(raw_chunks),
                        "upload_date": upload_date,
                    },
                }
            )

        # Add to vectorstore
        try:
            added = rag.add_chunks(chunks)
        except Exception as e:
            print(f"  [ERROR] add_chunks failed: {e}")
            stats["error"] += 1
            continue

        # Record in DB
        doc_record = {
            "doc_id": doc_id,
            "doc_type": doc_type,
            "title": title,
            "source_file": source_rel,
            "device_vendor": "",
            "device_type": "",
            "chunk_count": added,
            "upload_date": upload_date,
            "file_hash": file_hash,
        }
        db.upsert_document(doc_record)
        db.log_upload(doc_id, "upload", "success", f"Batch ingest: {added} chunks")

        stats["success"] += 1
        stats["total_chunks"] += added
        print(f"  [OK] {added} chunks ingested (doc_id={doc_id[:8]}...)")

    # Summary
    print(f"\n{'=' * 60}")
    print(f"Batch ingest complete:")
    print(f"  Success:  {stats['success']}")
    print(f"  Skipped:  {stats['skipped']} (hash unchanged)")
    print(f"  Empty:    {stats['empty']} (no content)")
    print(f"  Error:    {stats['error']}")
    print(f"  Total chunks: {stats['total_chunks']}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
